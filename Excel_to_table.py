from openpyxl import load_workbook
import os 
import re 

def create_redmine_mkdir(redmine):
    path_DB = '.\\{redmine}_DB'.format(redmine = redmine)
    path_BKDB = '.\\{redmine}_BKDB'.format(redmine = redmine)
    os.makedirs(path_DB, exist_ok=True)
    os.makedirs(path_BKDB, exist_ok=True)

def list_excel_files(dirpath):
    list_all_files = os.listdir(dirpath)
    excel_files = []
    for file in list_all_files:
        if '.xlsx' in file:
            excel_files.append(file)
    return excel_files

def excel_sheetnames_locate(excel_obj):
    book_sheet_names = excel_obj.sheetnames
    excel_file_in_list = []
    regex = re.compile(r"附件1-\d\Z")
    for sheet in book_sheet_names:
        match = regex.search(sheet)
        if match is not None:
        #    print(match.group(0))
            excel_file_in_list.append(match.group(0))
    return excel_file_in_list

class write_output_file():
    def __init__(self, sheet_obj, input_author, input_Issue):
        self.sheet_obj               = sheet_obj
        self.author                  = input_author
        self.input_Issue             = input_Issue
        self.table_name              = self.sheet_obj['B4'].value
        self.table_comment           = self.sheet_obj['B5'].value
        self.table_schema            = self.sheet_obj['B1'].value
        self.Table_label             = self.sheet_obj['G2'].value
        self.front_label_establish   = self.Table_label.split('/')[0]
        self.back_label_establish    = self.Table_label.split('/')[1]
        self.Table_Space_front_value = self.judge_front_label()
        self.Table_Space_back_value  = self.judge_back_label()

    def judge_front_label(self):
        if self.front_label_establish == 'Y':
            return self.judge_table_space_Value()

    def judge_back_label(self):
        if self.back_label_establish == 'Y':
            return self.judge_table_space_Value(BAT = 'BAT_')

    def judge_table_space_Value(self, BAT = ''):
        if 'PRO' in self.table_name:
            return 'TS_WMS_{BAT}PRO'.format(BAT = BAT)
        elif 'CUS' in self.table_name:
            return 'TS_WMS_{BAT}CUS'.format(BAT = BAT)
        elif 'ADM' in self.table_name:
            return 'TS_WMS_{BAT}ADM'.format(BAT = BAT)
        else:
            return 'TS_WMS_{BAT}COMMON'.format(BAT = BAT)

    def output_script_part1(self, input_author, input_Issue,DB):
        DB = DB
        output_script_part1 = """--===================================================================\n--AUTHOR      : {author}\n--COMMENT     : {Issue} {table_name}  {comment} {front_or_back}\n--===================================================================\nCREATE  TABLE WMSR6USR.{table_name} (\n """.format(author = input_author,Issue = input_Issue,table_name =self.table_name,comment =self.table_comment,front_or_back = '(前台)' if DB =='DB' else '(後台)')
        return output_script_part1
    
    def set_col_value(self, col_pre, col_remove_Add):
            col_remove = '{col_pre}{col_remove}'.format(col_pre = col_pre, col_remove = col_remove_Add)
            col_value  = self.sheet_obj[col_remove].value
            return col_value

    def output_script_part2(self, DB):
        col_all_list          = []
        col_Remark_Value_list = []
        output_script_part2   = []
        Pk_col_name_value     = []
        col_locate = 7
        DB = DB

        while self.set_col_value('A', col_remove_Add=col_locate):
            col_Name_Value    = self.set_col_value('A',col_remove_Add=col_locate)
            col_Type_Value    = self.set_col_value('B',col_remove_Add=col_locate)
            col_Null_Value    = self.set_col_value('E',col_remove_Add=col_locate)
            col_Default_Value = self.set_col_value('F',col_remove_Add=col_locate)
            col_Pk_Value      = self.set_col_value('C',col_remove_Add=col_locate)
            col_Remark_Value  = self.set_col_value('G',col_remove_Add=col_locate)

            col_all_list.append([col_Name_Value, col_Type_Value, col_Null_Value, col_Default_Value])
            col_Remark_Value_list.append([col_Name_Value, col_Remark_Value])
            if col_Pk_Value:
                Pk_col_name_value.append(col_Name_Value)
            col_locate +=1

        for row in col_all_list:
            col_value = (' 	{col_name} {col_type} {Set_null} {Set_default} {last_word}\n '.format(col_name = row[0],col_type = row[1],Set_null = '' if row[2] == None else row[2] 
            ,Set_default = '' if row[3] == None else ('DEFAULT {Default}').format(Default = row[3]),last_word = '' if row == col_all_list[-1] else ','  ))
            output_script_part2.append(col_value)
        
        col_index_value   = """	) COMPRESS YES ADAPTIVE IN {Table_space} INDEX IN TS_WMS{BAT}_IDX ORGANIZE BY ROW  @\n\n""".format(Table_space =  self.Table_Space_front_value if DB =='DB' else self.Table_Space_back_value,BAT ='' if DB=='DB' else '_BAT' )
        output_script_part2.append(col_index_value)

        if Pk_col_name_value:
            Pk_col_name_value = map(( lambda x: '\"' + x + '\"'), Pk_col_name_value)
            col_Pk_Value = """ALTER TABLE {Table_schema}.{Table_name} ADD CONSTRANT PK_{Table_name} PRIMARY KEY({Pk_col_name_value}) @\n""".format(Table_schema = self.table_schema,Table_name =self.table_name,Pk_col_name_value = ",".join(Pk_col_name_value))
            output_script_part2.append(col_Pk_Value)

        col_Remark_table_value = """COMMENT ON TABLE \"{Table_schema}\".\"{Table_name}\" IS\'{Table_comment}\'@\n""".format(Table_schema =self.table_schema,Table_name = self.table_name,Table_comment = self.table_comment)
        output_script_part2.append(col_Remark_table_value)

        for col_remark in col_Remark_Value_list:
            col_remark_value = """COMMENT ON COLUMN \"{Table_schema}\".\"{Table_name}\".\"{Table_col}\" IS \'{Table_remark}\'@\n""".format(Table_schema =self.table_schema,Table_name = self.table_name,Table_col = col_remark[0],Table_remark = col_remark[1])
            output_script_part2.append(col_remark_value)

        return "".join(output_script_part2)
    
    def find_locate_account(self):
        col_grant_locate = 1
        while self.set_col_value('A',col_grant_locate) != '帳號':
            col_grant_locate +=1 
            if self.set_col_value('A',col_grant_locate) == '帳號':
                return (col_grant_locate + 1)

    def output_script_part3(self):
        col_grant_locate = self.find_locate_account()
        grant_user_list       = []
        output_script_part3   = []

        while self.set_col_value('A', col_grant_locate):
            grant_permission_user    = self.set_col_value('A', col_remove_Add=col_grant_locate)
            grant_verb               = self.set_col_value('B', col_remove_Add=col_grant_locate)
            grant_user_list.append([grant_verb,grant_permission_user])
            col_grant_locate +=1
        
        for row in grant_user_list:
            grant_value = """\nGRANT {grant_verb} ON TABLE \"{table_schema}\".\"{table_name}\" TO USER \"{grant_permission_user}\"@""".format(grant_verb = row[0],table_schema =self.table_schema,table_name = self.table_name, grant_permission_user = row[1])
            output_script_part3.append(grant_value)
        return "".join(output_script_part3)


    def output_write_txt_file(self, DB):
        file_name = './{redmine}_{DB}/{table_name}.Table.sql'.format(table_name = self.table_name,redmine = self.input_Issue,DB = DB)
        with open(file=file_name, mode='w+', encoding='utf-8') as f :
            f.write(self.output_script_part1(input_author = self.author, input_Issue = self.input_Issue,DB=DB))
            f.write(self.output_script_part2(DB))
            f.write(self.output_script_part3())
            f.close()
    
    def start(self):
        if self.back_label_establish == 'Y':
            self.output_write_txt_file('BKDB')
        if self.front_label_establish == 'Y':
            self.output_write_txt_file('DB')

    
if __name__ == '__main__':
    input_dir_path = input('給我create TABLE，SDS位置?\n')
    input_author   = input('你哪位?\n')
    input_Issue    = input('單號是多少?\n')
    create_redmine_mkdir(input_Issue)
    excel_files = list_excel_files(input_dir_path)
    for excel_file in excel_files:
        book = load_workbook(excel_file)
        print('正在處理:{excel_file}'.format(excel_file = excel_file))
        sheet_list = excel_sheetnames_locate(excel_obj = book)
        for sheet in sheet_list :
            write_output_file_obj = write_output_file(sheet_obj=book[sheet], input_author=input_author, input_Issue=input_Issue)
            write_output_file_obj.start()
            
